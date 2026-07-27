import json
import logging
from datetime import datetime, timezone
from typing import AsyncGenerator, Optional

from dotenv import load_dotenv
load_dotenv()

from fastapi import FastAPI, File, Form, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, RedirectResponse, Response, StreamingResponse

from app.agent_client import (
    request_chart_from_agent, research_agent, format_validated, classify_card_type, doc_agent,
)
from app.firestore_client import (
    save_graph, find_cached_card, get_stored_card, list_published_cards,
)
from app.quota import uid_from_request, client_ip, consume_quota, is_internal
from app.validator import validate_card
from app.schemas import WeirdCard, GenerateRequest
from app import seo

logger = logging.getLogger("uvicorn.error")

app = FastAPI(title="WeirdStats Backend")

import os

# Local dev + hosted frontend origins. Add extras via CORS_ORIGINS (comma-separated).
_default_origins = [
    "http://localhost:4200", "http://localhost:8100",
    "http://localhost:8080", "capacitor://localhost", "ionic://localhost",
    # Custom domain + new project hosting
    "https://weirdstats.ai", "https://www.weirdstats.ai",
    "https://weirdstats-ai.web.app", "https://weirdstats-ai.firebaseapp.com",
    # Legacy project (kept until fully retired)
    "https://weirdstatsai-aaaf7.web.app",
    "https://weirdstatsai-aaaf7.firebaseapp.com",
]
_extra = [o.strip() for o in os.getenv("CORS_ORIGINS", "").split(",") if o.strip()]

app.add_middleware(
    CORSMiddleware,
    allow_origins=_default_origins + _extra,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Stripe billing (checkout / customer portal / webhook). Self-contained router;
# billing endpoints return 503 until the STRIPE_* env vars are configured.
from app.billing import router as billing_router
app.include_router(billing_router)


@app.get("/health")
async def health() -> dict:
    return {"status": "ok"}


# ── Business / enterprise contact form ──────────────────────────────────────
from pydantic import BaseModel
from fastapi import HTTPException
from fastapi.concurrency import run_in_threadpool


class ContactRequest(BaseModel):
    name: str = ""
    email: str = ""
    company: str = ""
    companySize: str = ""
    role: str = ""
    phone: str = ""
    topic: str = ""
    message: str = ""
    website: str = ""   # honeypot — bots fill hidden fields; humans never see it


def _send_contact_email(d: dict) -> bool:
    """Deliver a contact submission to the team inbox via SMTP. Best-effort:
    returns False (and logs) when SMTP isn't configured, so the endpoint still
    succeeds and the lead is preserved in Firestore. Reply-To is the customer's
    address, so a plain 'Reply' reaches them directly."""
    import smtplib
    import ssl
    from email.message import EmailMessage

    host = os.getenv("SMTP_HOST", "smtp.gmail.com")
    port = int(os.getenv("SMTP_PORT", "587"))
    user = os.getenv("SMTP_USER", "")
    pwd = os.getenv("SMTP_PASSWORD", "")
    to = os.getenv("CONTACT_TO", "weirdstats.ai@gmail.com")
    if not user or not pwd:
        logger.warning("Contact email skipped: SMTP_USER/SMTP_PASSWORD not set")
        return False

    msg = EmailMessage()
    who = d.get("company") or d.get("name") or "New inquiry"
    msg["Subject"] = f"[WeirdStats Business] {d.get('topic') or 'Inquiry'} — {who}"
    msg["From"] = f"WeirdStats Contact <{user}>"
    msg["To"] = to
    if d.get("email"):
        msg["Reply-To"] = d["email"]
    msg.set_content(
        "New business inquiry from the WeirdStats contact form.\n\n"
        f"Name:          {d.get('name','')}\n"
        f"Work email:    {d.get('email','')}\n"
        f"Company:       {d.get('company','')}\n"
        f"Company size:  {d.get('companySize','')}\n"
        f"Role:          {d.get('role','')}\n"
        f"Phone:         {d.get('phone','')}\n"
        f"Interested in: {d.get('topic','')}\n\n"
        f"Message:\n{d.get('message','')}\n"
    )
    ctx = ssl.create_default_context()
    with smtplib.SMTP(host, port, timeout=20) as s:
        s.starttls(context=ctx)
        s.login(user, pwd)
        s.send_message(msg)
    return True


@app.post("/api/contact")
async def contact(req: ContactRequest) -> dict:
    # Honeypot tripped → almost certainly a bot. Pretend success, do nothing.
    if req.website.strip():
        return {"ok": True}

    name = req.name.strip()
    email = req.email.strip()
    message = req.message.strip()
    if not name or "@" not in email or "." not in email.split("@")[-1] or len(message) < 10:
        raise HTTPException(status_code=422, detail="Please provide a name, a valid work email, and a message.")
    if len(message) > 5000 or len(name) > 200 or len(email) > 200:
        raise HTTPException(status_code=422, detail="One of the fields is too long.")

    data = {
        "name": name, "email": email,
        "company": req.company.strip()[:200], "companySize": req.companySize.strip()[:40],
        "role": req.role.strip()[:120], "phone": req.phone.strip()[:60],
        "topic": req.topic.strip()[:80], "message": message,
    }

    # Preserve the lead first — email delivery is best-effort on top of this.
    try:
        from app.firestore_client import _get_db
        _get_db().collection("contactMessages").add({
            **data,
            "source": "contact_form",
            "handled": False,
            "createdAt": datetime.now(timezone.utc).isoformat(),
        })
    except Exception:
        logger.warning("Failed to store contact message", exc_info=True)

    emailed = False
    try:
        emailed = await run_in_threadpool(_send_contact_email, data)
    except Exception:
        logger.warning("Failed to send contact email", exc_info=True)

    return {"ok": True, "emailed": emailed}


@app.get("/api/admin/trending")
async def admin_trending(geo: str = "US") -> dict:
    """Trending topics for the admin panel to turn into Home-feed cards.

    Pulls from Google Trends (daily trending searches) plus Google News top
    stories and a politics query — all public RSS. Returns a deduped list
    tagged by source so the panel can group them. Fails soft per source.
    """
    import xml.etree.ElementTree as ET
    import httpx

    sources = [
        ("trends",   f"https://trends.google.com/trending/rss?geo={geo}"),
        ("news",     f"https://news.google.com/rss?hl=en-US&gl={geo}&ceid={geo}:en"),
        ("politics", f"https://news.google.com/rss/search?q=politics&hl=en-US&gl={geo}&ceid={geo}:en"),
    ]
    per_source = 12
    topics: list[dict] = []
    seen: set[str] = set()

    async with httpx.AsyncClient(
        timeout=10.0, follow_redirects=True,
        headers={"User-Agent": "Mozilla/5.0 (compatible; WeirdStatsBot/1.0)"},
    ) as client:
        for source, url in sources:
            try:
                resp = await client.get(url)
                root = ET.fromstring(resp.text)
                count = 0
                for item in root.iter("item"):
                    title = (item.findtext("title") or "").strip()
                    # Google News titles carry a " - Publisher" suffix; drop it.
                    if source != "trends" and " - " in title:
                        title = title.rsplit(" - ", 1)[0].strip()
                    if not title:
                        continue
                    key = title.lower()
                    if key in seen:
                        continue
                    seen.add(key)
                    topics.append({"topic": title, "source": source})
                    count += 1
                    if count >= per_source:
                        break
            except Exception:
                logger.warning("trending source failed: %s", source, exc_info=True)

    return {"topics": topics, "geo": geo}


# ── SEO: bot-snapshot rendering for shareable card URLs ─────────────────────
# Firebase Hosting rewrites /card/**, /share/**, /og/**, /sitemap-cards.xml to
# this service. Bots get purpose-built HTML/images; humans get the SPA shell.

def _card_is_public(doc) -> bool:
    """A card may be rendered to the public (bot snapshot / OG image) only when
    it is published or curated onto a public feed — mirrors the public-read
    conditions in firestore.rules. Private/draft cards must never leak through
    these routes even though the Admin SDK bypasses those rules."""
    if not doc:
        return False
    return (
        doc.get("publishStatus") == "published"
        or doc.get("homeFeatured") is True
        or doc.get("showOnHome") is True
        or doc.get("showOnExplore") is True
    )


async def _card_route(card_id: str, request: Request) -> Response:
    # These responses branch on User-Agent (bot snapshot vs SPA shell). Firebase's
    # CDN caches by URL, NOT by UA, so a cached bot snapshot would be served to
    # humans (and vice-versa). Force `no-store` so every request reaches Cloud Run
    # and gets UA-correct content. The og-image + sitemap routes below don't
    # branch on UA, so they stay publicly cacheable.
    no_store = {"Cache-Control": "private, no-store"}
    ua = request.headers.get("user-agent", "")
    if seo.is_bot(ua):
        doc = get_stored_card(card_id)
        # Never leak private/draft content to scrapers — treat it as not-found.
        if not _card_is_public(doc):
            doc = None
        html = seo.build_snapshot_html(card_id, doc)
        return HTMLResponse(html, headers=no_store)
    shell = await seo.get_spa_shell()
    return HTMLResponse(shell, headers=no_store)


@app.get("/card/{card_id}")
async def card_page(card_id: str, request: Request) -> Response:
    return await _card_route(card_id, request)


@app.get("/share/{card_id}")
async def share_page(card_id: str, request: Request) -> Response:
    return await _card_route(card_id, request)


@app.get("/og/card/{ref}")
async def og_card_image(ref: str) -> Response:
    card_id = ref[:-4] if ref.endswith(".png") else ref
    doc = get_stored_card(card_id)
    # Only published / feed-curated cards get a real preview image; private or
    # draft cards fall back to the generic OG image instead of leaking content.
    if not _card_is_public(doc):
        return RedirectResponse(seo.DEFAULT_OG_IMAGE, status_code=307)
    png = seo.compose_og_image(doc)
    if png is None:
        return RedirectResponse(seo.DEFAULT_OG_IMAGE, status_code=307)
    # Short fresh window so card edits / renderer fixes propagate within the
    # hour; stale-while-revalidate keeps it instant for scrapers meanwhile.
    return Response(png, media_type="image/png",
                    headers={"Cache-Control": "public, max-age=3600, stale-while-revalidate=86400"})


@app.get("/sitemap-cards.xml")
async def sitemap_cards() -> Response:
    xml = seo.build_cards_sitemap(list_published_cards())
    return Response(xml, media_type="application/xml",
                    headers={"Cache-Control": "public, max-age=3600"})


def _fallback_card(prompt: str, reason: str) -> dict:
    """Minimal valid card when the pipeline fails — keeps the app responsive."""
    return validate_card({
        "status": "needs_review",
        "title": prompt.strip()[:80] or "Couldn't generate a card",
        "cardType": "fact",
        "presentationType": "fact",
        "insight": f"We couldn't verify data for this one right now. ({reason})",
        "tags": ["stats"],
        "weirdScore": 3,
        "uiMeta": {"category": "Other", "icon": "🤔", "accentColor": "#6C5CE7"},
        "dataMeta": {"dataMode": "estimated", "confidence": "low"},
    })


def _sse(event_type: str, payload: dict) -> str:
    return f"data: {json.dumps({'type': event_type, **payload})}\n\n"


@app.post("/api/generate/stream")
async def generate_stream(req: GenerateRequest, request: Request) -> StreamingResponse:
    # Identity comes from a verified ID token — NEVER req.uid, which the caller
    # controls. Resolved here (outside the generator) so it happens before any
    # work is scheduled.
    caller = uid_from_request(request.headers.get("authorization"))
    caller_ip = client_ip(request.headers, request.client.host if request.client else None)

    async def stream() -> AsyncGenerator[str, None]:
        # Cache hit — return instantly, and don't bill it against the quota:
        # serving a stored card costs a Firestore read, not an OpenAI call.
        cached = find_cached_card(req.prompt)
        if cached:
            yield _sse("card", {"data": cached})
            return

        # Everything past here spends money, so meter it. Atomic check-and-take
        # in Firestore — the client-side cap is a UX hint, this is the gate.
        allowed, why = (True, "") if is_internal(request.headers) else \
            await run_in_threadpool(consume_quota, caller, caller_ip)
        if not allowed:
            yield _sse("limit", {"message": why})
            yield _sse("error", {"message": why})
            return

        # Step 1: research
        yield _sse("status", {"message": "Searching the web…", "step": 1})
        try:
            brief = await research_agent(req.prompt)
        except Exception as e:
            yield _sse("error", {"message": "Research failed. Try again."})
            return

        # Step 1.5: dedicated cardType decision (fails open — None lets the
        # Format Agent decide, so a classifier hiccup never blocks a card).
        card_type = await classify_card_type(req.prompt, brief)

        # Tell the client the card shape the moment we know it — the UI renders a
        # matching skeleton while the (slower) format step runs, so the wait feels
        # shorter. Fails open: if the classifier returned None, no shape is sent
        # and the client just keeps its generic loading state.
        if card_type:
            yield _sse("shape", {"cardType": card_type})

        # Step 2: format — retry once before surfacing an error, since the
        # format/validate step occasionally fails transiently on the first try.
        yield _sse("status", {"message": "Building your card…", "step": 2})
        # format_validated formats, validates, and guarantees the card has data
        # its type can actually render (repairs a hollow card, degrading as a last
        # resort). The try/except here only guards transient API failures.
        try:
            card = await format_validated(brief, card_type)
        except Exception:
            logger.warning("Stream format failed, retrying once", exc_info=True)
            try:
                card = await format_validated(brief, card_type)
            except Exception:
                logger.warning("Stream format failed twice", exc_info=True)
                yield _sse("error", {"message": "Could not format card. Try again."})
                return

        # Step 3: save a cache-only copy (not user-owned) + return.
        # Drafts live on the device; a card only enters a user's collection
        # when they explicitly Save publicly/privately from the app.
        yield _sse("status", {"message": "Almost done…", "step": 3})
        graph_id = save_graph(card, req.prompt, uid=None)
        result = {
            **card,
            "id": graph_id,
            "prompt": req.prompt,
            "createdAt": datetime.now(timezone.utc).isoformat(),
        }
        yield _sse("card", {"data": result})

    return StreamingResponse(
        stream(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
        },
    )


MAX_PDF_BYTES = int(os.getenv("MAX_PDF_BYTES", str(15 * 1024 * 1024)))  # 15 MB

# Text-based formats are extracted server-side and sent to the agent as text;
# PDFs go natively (vision) so tables/charts/scans work.
TEXT_EXTS = {"csv", "tsv", "txt", "md"}
SUPPORTED_EXTS = {"pdf", "docx", "xlsx"} | TEXT_EXTS


def _file_ext(filename: str) -> str:
    return filename.rsplit(".", 1)[-1].lower() if "." in filename else ""


def _docx_text(data: bytes) -> str:
    """Paragraph/table text from a .docx (a zip of XML) — stdlib only."""
    import io
    import re as _re
    import zipfile
    with zipfile.ZipFile(io.BytesIO(data)) as z:
        xml = z.read("word/document.xml").decode("utf-8", errors="replace")
    # Cell/paragraph boundaries → tabs/newlines so table structure survives.
    xml = _re.sub(r"</w:tc>", "\t", xml)
    xml = _re.sub(r"</w:p>|</w:tr>", "\n", xml)
    text = _re.sub(r"<[^>]+>", "", xml)
    return "\n".join(line.strip() for line in text.splitlines() if line.strip())


def _xlsx_text(data: bytes, max_rows_per_sheet: int = 500) -> str:
    """Sheets flattened to TSV-ish text via openpyxl."""
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    parts: list[str] = []
    for ws in wb.worksheets:
        parts.append(f"## Sheet: {ws.title}")
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows_per_sheet:
                parts.append(f"... (truncated at {max_rows_per_sheet} rows)")
                break
            cells = ["" if c is None else str(c) for c in row]
            if any(cells):
                parts.append("\t".join(cells))
    return "\n".join(parts)


def _extract_doc_text(ext: str, data: bytes) -> str:
    if ext == "docx":
        return _docx_text(data)
    if ext == "xlsx":
        return _xlsx_text(data)
    return data.decode("utf-8", errors="replace")


MAX_PDF_PAGES = int(os.getenv("MAX_PDF_PAGES", "40"))


def _slice_pdf(data: bytes, max_pages: int) -> tuple[bytes, int, int]:
    """Return (pdf_bytes, total_pages, used_pages). Long documents are cut to
    the first `max_pages` so they fit the model's context window — a 300-page
    annual report otherwise 400s with context_length_exceeded. If pypdf can't
    parse the file, fall back to the original bytes and let the model try."""
    import io
    try:
        from pypdf import PdfReader, PdfWriter
        reader = PdfReader(io.BytesIO(data))
        total = len(reader.pages)
        if total <= max_pages:
            return data, total, total
        writer = PdfWriter()
        for page in reader.pages[:max_pages]:
            writer.add_page(page)
        out = io.BytesIO()
        writer.write(out)
        return out.getvalue(), total, max_pages
    except Exception:
        logger.warning("PDF slicing failed — sending original file", exc_info=True)
        return data, 0, 0


@app.post("/api/projects/import/stream")
async def project_import_stream(
    request: Request,
    file: UploadFile = File(...),
    max_findings: Optional[int] = Form(None),
) -> StreamingResponse:
    """Bulk import: read a document (PDF, Word, Excel, CSV, text), extract
    stat-worthy findings, and stream one validated WeirdCard per finding.
    `max_findings` lets the client cap extraction to the project's remaining
    space (never raises the server-side ceiling). Pipeline per finding mirrors
    the single-prompt flow: classify -> format -> validate. The frontend owns
    persistence (it saves each card into the project with the user's auth
    context)."""
    file_bytes = await file.read()
    filename = file.filename or "document"
    ext = _file_ext(filename)
    cap = max(1, min(max_findings, 30)) if max_findings else None

    async def stream() -> AsyncGenerator[str, None]:
        if not file_bytes:
            yield _sse("error", {"message": "Empty file."})
            return
        if len(file_bytes) > MAX_PDF_BYTES:
            yield _sse("error", {"message": "File is too large (max 15 MB)."})
            return
        if ext not in SUPPORTED_EXTS:
            yield _sse("error", {"message": "Unsupported file type. Use PDF, Word (.docx), Excel (.xlsx), CSV, or TXT."})
            return

        yield _sse("status", {"message": "Reading your document…", "step": 1})
        try:
            if ext == "pdf":
                pdf_bytes, total_pages, used_pages = _slice_pdf(file_bytes, MAX_PDF_PAGES)
                if total_pages > used_pages > 0:
                    yield _sse("status", {
                        "message": f"Long document — reading the first {used_pages} of {total_pages} pages…",
                        "step": 1,
                    })
                extraction = await doc_agent(filename, pdf_bytes=pdf_bytes, max_findings=cap)
            else:
                doc_text = _extract_doc_text(ext, file_bytes)
                if not doc_text.strip():
                    yield _sse("error", {"message": "Could not read any text from this file."})
                    return
                extraction = await doc_agent(filename, text=doc_text, max_findings=cap)
        except Exception as e:
            logger.warning("doc_agent failed", exc_info=True)
            err = str(e)
            if "insufficient_quota" in err or "429" in err:
                msg = "The AI service is out of capacity right now. Please try again later."
            elif "context_length" in err or "context window" in err:
                msg = "This document is too long to read in one go. Try a shorter file or export just the key pages."
            else:
                msg = "Could not read this document. Try another file."
            yield _sse("error", {"message": msg})
            return

        findings = extraction.get("findings", [])
        if not findings:
            yield _sse("error", {"message": "No chartable data found in this document."})
            return

        total = len(findings)
        yield _sse("plan", {
            "documentTitle": extraction.get("documentTitle", filename),
            "total": total,
        })

        built = 0
        for i, finding in enumerate(findings, start=1):
            question = finding["question"]
            brief = finding["brief"]
            yield _sse("status", {
                "message": f"Building card {i} of {total}…",
                "step": i + 1,
                "question": question,
            })
            try:
                card_type = await classify_card_type(question, brief)
                card = await format_validated(brief, card_type)
            except Exception:
                # One bad finding never sinks the import — skip and continue.
                logger.warning("Import card %d/%d failed (%s)", i, total, question, exc_info=True)
                yield _sse("skipped", {"question": question, "index": i, "total": total})
                continue

            built += 1
            graph_id = save_graph(card, question, uid=None)
            yield _sse("card", {
                "data": {
                    **card,
                    "id": graph_id,
                    "prompt": question,
                    "createdAt": datetime.now(timezone.utc).isoformat(),
                },
                "index": i,
                "total": total,
            })

        yield _sse("done", {"built": built, "total": total})

    return StreamingResponse(
        stream(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
        },
    )


@app.post("/api/generate", response_model=WeirdCard)
async def generate(req: GenerateRequest, request: Request) -> dict:
    # 1. Dedup — serve an existing card for this question if we have one.
    cached = find_cached_card(req.prompt)
    if cached:
        return cached

    # 2. Same gate as the streaming endpoint: verified identity, metered spend.
    caller = uid_from_request(request.headers.get("authorization"))
    allowed, why = (True, "") if is_internal(request.headers) else \
        await run_in_threadpool(
            consume_quota, caller,
            client_ip(request.headers, request.client.host if request.client else None))
    if not allowed:
        raise HTTPException(status_code=429, detail=why)

    # 2. Run the two-step pipeline, validate. Retry once, then fall back.
    card: dict
    try:
        raw = await request_chart_from_agent(req.prompt, req.preferredType)
        card = validate_card(raw)
    except Exception:
        logger.warning("Pipeline failed (attempt 1), retrying once", exc_info=True)
        try:
            raw = await request_chart_from_agent(req.prompt, req.preferredType)
            card = validate_card(raw)
        except Exception as e:
            logger.warning("Pipeline failed (attempt 2), using fallback", exc_info=True)
            card = _fallback_card(req.prompt, str(e)[:80])

    # 3. Persist a cache-only copy (not user-owned) and return.
    graph_id = save_graph(card, req.prompt, uid=None)
    return {
        **card,
        "id": graph_id,
        "prompt": req.prompt,
        "createdAt": datetime.now(timezone.utc).isoformat(),
    }
